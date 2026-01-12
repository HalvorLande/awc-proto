from __future__ import annotations

import os
import re
import shutil
from datetime import datetime, timezone, date
from pathlib import Path
from typing import Any, Dict, Iterable, Optional, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook
from urllib.parse import quote_plus

from sqlalchemy import create_engine, text


# ------------------------------------------------------------
# Config
# ------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parents[3]  # backend/app/jobs -> repo root
IMPORT_DIR = REPO_ROOT / "import"
IMPORTED_DIR = IMPORT_DIR / "imported"

SHEET_FIRMAINFO = "Proff Forvalt - Firmainfo"
SHEET_CONTACTS = "Proff Forvalt - Kontaktpersoner"

SOURCE_NAME = "proff_forvalt_excel"
ACCOUNT_VIEW = "company"  # these exports are typically company-level unless stated otherwise


# ------------------------------------------------------------
# Load env
# ------------------------------------------------------------
load_dotenv(dotenv_path=REPO_ROOT / "backend" / ".env", override=False)


# ------------------------------------------------------------
# DB engine (env vars)
# ------------------------------------------------------------
def make_engine():
    server = os.getenv("SQL_SERVER", "AAD-GM12FD8W")
    database = os.getenv("SQL_DATABASE", "AwcProto")
    driver = os.getenv("SQL_DRIVER", "ODBC Driver 17 for SQL Server")

    odbc_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        "Trusted_Connection=yes;"
        "TrustServerCertificate=yes;"
    )
    url = "mssql+pyodbc:///?odbc_connect=" + quote_plus(odbc_str)
    return create_engine(url, future=True, pool_pre_ping=True)


# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------
NBSP = "\u00A0"

def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def normalize_orgnr(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = "".join(ch for ch in str(v) if ch.isdigit())
    return s if len(s) == 9 else None

def parse_number(v: Any) -> Optional[float]:
    """
    Handles:
      - 123 456
      - 6 499,00 (NBSP + comma decimal)
      - empty cells
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(NBSP, "").replace(" ", "")
    # Norwegian decimal comma
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    # Remove any stray non-numeric chars (except minus and dot)
    s = re.sub(r"[^0-9\.\-]", "", s)
    if s in ("", "-", ".", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None

def parse_int(v: Any) -> Optional[int]:
    n = parse_number(v)
    if n is None:
        return None
    try:
        return int(round(n))
    except Exception:
        return None

def parse_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None

    # Supports "dd.mm.yyyy" and "yyyy-mm-dd"
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def read_sheet_as_rows(ws) -> Tuple[list[str], list[dict[str, Any]]]:
    """
    Returns (headers, rows), where each row is dict header->cell_value.
    """
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    out_rows: list[dict[str, Any]] = []
    for r in rows_iter:
        if r is None:
            continue
        row_dict = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        # skip fully empty lines
        if all(v is None or str(v).strip() == "" for v in row_dict.values()):
            continue
        out_rows.append(row_dict)
    return headers, out_rows


# ------------------------------------------------------------
# SQL: ensure contact table exists
# ------------------------------------------------------------
ENSURE_CONTACT_TABLE = """
IF OBJECT_ID('dbo.company_contact_person', 'U') IS NULL
BEGIN
    CREATE TABLE dbo.company_contact_person (
        id INT IDENTITY(1,1) NOT NULL PRIMARY KEY,
        orgnr VARCHAR(9) NOT NULL,
        company_name NVARCHAR(255) NULL,
        person_name NVARCHAR(255) NOT NULL,
        role NVARCHAR(100) NULL,
        started_date DATE NULL,
        phone NVARCHAR(50) NULL,
        email NVARCHAR(255) NULL,
        postal_address NVARCHAR(255) NULL,
        postal_postnr NVARCHAR(20) NULL,
        postal_city NVARCHAR(100) NULL,
        business_address NVARCHAR(255) NULL,
        business_postnr NVARCHAR(20) NULL,
        business_city NVARCHAR(100) NULL,
        revenue FLOAT NULL,
        employees INT NULL,
        source_file NVARCHAR(260) NULL,
        imported_at_utc DATETIME2(0) NOT NULL CONSTRAINT DF_contact_imported DEFAULT SYSUTCDATETIME()
    );

    CREATE UNIQUE INDEX UX_contact_unique
    ON dbo.company_contact_person(orgnr, person_name, [role], started_date);
END;
"""


# ------------------------------------------------------------
# SQL MERGE statements
# (Assumes dbo.company columns exist: orgnr, name, phone, email, website, municipality, nace, etc.)
# If any column doesn't exist in your dbo.company, remove it here.
# ------------------------------------------------------------
MERGE_COMPANY = """
MERGE dbo.company WITH (HOLDLOCK) AS tgt
USING (SELECT
    :orgnr AS orgnr,
    :name AS name,
    :phone AS phone,
    :email AS email,
    :website AS website,
    :municipality AS municipality,
    :nace AS nace,
    :nace_desc AS nace_desc,
    :orgform AS orgform,
    :status AS status,
    SYSUTCDATETIME() AS updated_at
) AS src
ON tgt.orgnr = src.orgnr
WHEN MATCHED THEN UPDATE SET
    name = COALESCE(src.name, tgt.name),
    phone = COALESCE(src.phone, tgt.phone),
    email = COALESCE(src.email, tgt.email),
    website = COALESCE(src.website, tgt.website),
    municipality = COALESCE(src.municipality, tgt.municipality),
    nace = COALESCE(src.nace, tgt.nace),
    updated_at = src.updated_at
WHEN NOT MATCHED THEN
    INSERT (orgnr, name, phone, email, website, municipality, nace, created_at, updated_at)
    VALUES (src.orgnr, src.name, src.phone, src.email, src.website, src.municipality, src.nace, SYSUTCDATETIME(), src.updated_at);
"""

# Financial statement: your table schema uses (orgnr, year, account_view) conceptually.
# We'll MERGE by (orgnr, year, account_view) by matching with ISNULL(account_view,'company').
MERGE_FIN_STATEMENT = """
MERGE dbo.financial_statement WITH (HOLDLOCK) AS tgt
USING (SELECT
    :orgnr AS orgnr,
    :year AS [year],
    :account_view AS account_view,
    :revenue AS revenue,
    :ebitda AS ebitda,
    :ebit AS ebit,
    :assets AS assets,
    :equity AS equity,
    :cash AS cash,
    :source AS source,
    :fetched_at_utc AS fetched_at_utc
) AS src
ON tgt.orgnr = src.orgnr
AND tgt.[year] = src.[year]
AND ISNULL(tgt.account_view, 'company') = ISNULL(src.account_view, 'company')
WHEN MATCHED THEN UPDATE SET
    revenue = COALESCE(src.revenue, tgt.revenue),
    ebitda = COALESCE(src.ebitda, tgt.ebitda),
    ebit = COALESCE(src.ebit, tgt.ebit),
    assets = COALESCE(src.assets, tgt.assets),
    equity = COALESCE(src.equity, tgt.equity),
    source = src.source,
    fetched_at_utc = src.fetched_at_utc,
    account_view = COALESCE(tgt.account_view, src.account_view, 'company')
WHEN NOT MATCHED THEN
    INSERT (orgnr, [year], revenue, ebitda, ebit, assets, equity, source, fetched_at_utc, account_view)
    VALUES (src.orgnr, src.[year], src.revenue, src.ebitda, src.ebit, src.assets, src.equity, src.source, src.fetched_at_utc, src.account_view);
"""

# Optional: also upsert selected codes into proff_financial_item (if you want to keep that fact table aligned)
MERGE_PROFF_FIN_ITEM = """
MERGE dbo.proff_financial_item WITH (HOLDLOCK) AS tgt
USING (SELECT
    :orgnr AS orgnr,
    :fiscal_year AS fiscal_year,
    :account_view AS account_view,
    :code AS code,
    :value AS value,
    :currency AS currency,
    :unit AS unit,
    :fetched_at_utc AS fetched_at_utc,
    :source AS source
) AS src
ON tgt.orgnr = src.orgnr
AND tgt.fiscal_year = src.fiscal_year
AND tgt.account_view = src.account_view
AND tgt.code = src.code
WHEN MATCHED THEN UPDATE SET
    value = src.value,
    currency = COALESCE(src.currency, tgt.currency),
    unit = COALESCE(src.unit, tgt.unit),
    fetched_at_utc = src.fetched_at_utc,
    source = src.source
WHEN NOT MATCHED THEN
    INSERT (orgnr, fiscal_year, account_view, code, value, currency, unit, fetched_at_utc, source)
    VALUES (src.orgnr, src.fiscal_year, src.account_view, src.code, src.value, src.currency, src.unit, src.fetched_at_utc, src.source);
"""

MERGE_CONTACT = """
MERGE dbo.company_contact_person WITH (HOLDLOCK) AS tgt
USING (SELECT
    :orgnr AS orgnr,
    :company_name AS company_name,
    :person_name AS person_name,
    :role AS role,
    TRY_CONVERT(date, :started_date) AS started_date,
    :phone AS phone,
    :email AS email,
    :postal_address AS postal_address,
    :postal_postnr AS postal_postnr,
    :postal_city AS postal_city,
    :business_address AS business_address,
    :business_postnr AS business_postnr,
    :business_city AS business_city,
    :revenue AS revenue,
    :employees AS employees,
    :source_file AS source_file,
    SYSUTCDATETIME() AS imported_at_utc
) AS src
ON tgt.orgnr = src.orgnr
AND tgt.person_name = src.person_name
AND ISNULL(tgt.role,'') = ISNULL(src.role,'')
AND ISNULL(tgt.started_date, '1900-01-01') = ISNULL(src.started_date, '1900-01-01')
WHEN MATCHED THEN UPDATE SET
    company_name = COALESCE(src.company_name, tgt.company_name),
    phone = COALESCE(src.phone, tgt.phone),
    email = COALESCE(src.email, tgt.email),
    postal_address = COALESCE(src.postal_address, tgt.postal_address),
    postal_postnr = COALESCE(src.postal_postnr, tgt.postal_postnr),
    postal_city = COALESCE(src.postal_city, tgt.postal_city),
    business_address = COALESCE(src.business_address, tgt.business_address),
    business_postnr = COALESCE(src.business_postnr, tgt.business_postnr),
    business_city = COALESCE(src.business_city, tgt.business_city),
    revenue = COALESCE(src.revenue, tgt.revenue),
    employees = COALESCE(src.employees, tgt.employees),
    source_file = COALESCE(src.source_file, tgt.source_file),
    imported_at_utc = src.imported_at_utc
WHEN NOT MATCHED THEN
    INSERT (
        orgnr, company_name, person_name, role, started_date, phone, email,
        postal_address, postal_postnr, postal_city,
        business_address, business_postnr, business_city,
        revenue, employees, source_file, imported_at_utc
    )
    VALUES (
        src.orgnr, src.company_name, src.person_name, src.role, src.started_date, src.phone, src.email,
        src.postal_address, src.postal_postnr, src.postal_city,
        src.business_address, src.business_postnr, src.business_city,
        src.revenue, src.employees, src.source_file, src.imported_at_utc
    );
"""


# ------------------------------------------------------------
# Column mapping: Firmainfo financial fields → normalized
# ------------------------------------------------------------
# We map columns by prefix and year suffix ", YYYY"
METRIC_PREFIX_TO_FIELD = {
    "Sum driftsinnt.": "revenue",
    "Driftsres.": "ebit",
    "Sum eiend.": "assets",
    "Sum egenkap.": "equity",
    "Kasse/bank/post": "cash",
    "Avskr. varige driftsmidl.": "depr",  # for computing EBITDA = EBIT + depreciation (if no EBITDA column)
    "Sum salgsinntekter": "sales_revenue",
}

# Optional mapping into proff_financial_item codes (core only)
FIELD_TO_PROFF_CODE = {
    "revenue": "SDI",
    "ebit": "DR",
    "assets": "SED",
    "equity": "SEK",
    "cash": "KBP",
}


def extract_year_metrics(row: dict[str, Any]) -> Dict[int, Dict[str, float]]:
    """
    Returns dict: {year: {field: value}}
    Looks for columns like "Driftsres., 2024" etc.
    """
    out: Dict[int, Dict[str, float]] = {}
    for k, v in row.items():
        if not k or v is None:
            continue
        k = str(k).strip()
        m = re.match(r"^(.*?),\s*(\d{4})$", k)
        if not m:
            continue
        prefix = m.group(1).strip()
        year = int(m.group(2))
        if prefix not in METRIC_PREFIX_TO_FIELD:
            continue
        field = METRIC_PREFIX_TO_FIELD[prefix]
        val = parse_number(v)
        if val is None:
            continue
        out.setdefault(year, {})[field] = val
    return out


def import_one_file(engine, path: Path) -> None:
    wb = load_workbook(path, data_only=True)

    if SHEET_FIRMAINFO not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET_FIRMAINFO}' in {path.name}")
    if SHEET_CONTACTS not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET_CONTACTS}' in {path.name}")

    ws1 = wb[SHEET_FIRMAINFO]
    ws2 = wb[SHEET_CONTACTS]

    _, firm_rows = read_sheet_as_rows(ws1)
    _, contact_rows = read_sheet_as_rows(ws2)

    fetched_at = now_utc().replace(tzinfo=None)  # store naive UTC in datetime2
    source_file = path.name

    with engine.begin() as conn:
        # Ensure contact table exists
        conn.execute(text(ENSURE_CONTACT_TABLE))

        # --- Sheet 1: Firmainfo ---
        for r in firm_rows:
            orgnr = normalize_orgnr(r.get("Orgnr"))
            if not orgnr:
                continue

            name = (r.get("Juridisk selskapsnavn") or r.get("Markedsnavn") or "").strip() if r.get("Juridisk selskapsnavn") else r.get("Markedsnavn")
            if isinstance(name, str):
                name = name.strip()
            else:
                name = None

            company_params = {
                "orgnr": orgnr,
                "name": name,
                "phone": (str(r.get("Telefon")).strip() if r.get("Telefon") else None),
                "email": (str(r.get("E-post")).strip() if r.get("E-post") else None),
                "website": (str(r.get("Internett")).strip() if r.get("Internett") else None),
                "municipality": (str(r.get("Kommune")).strip() if r.get("Kommune") else None),
                "nace": (str(r.get("NACE-bransjekode")).strip() if r.get("NACE-bransjekode") else None),
                "nace_desc": (str(r.get("NACE-beskrivelse")).strip() if r.get("NACE-beskrivelse") else None),
                "orgform": (str(r.get("Org.form")).strip() if r.get("Org.form") else None),
                "status": (str(r.get("Status")).strip() if r.get("Status") else None),
            }
            conn.execute(text(MERGE_COMPANY), company_params)

            # Financials by year
            ym = extract_year_metrics(r)
            for year, metrics in ym.items():
                revenue = metrics.get("revenue") or metrics.get("sales_revenue")
                ebit = metrics.get("ebit")
                assets = metrics.get("assets")
                equity = metrics.get("equity")
                cash = metrics.get("cash")
                depr = metrics.get("depr")

                # EBITDA: if explicitly exists in sheet later, you can map it;
                # otherwise approximate = EBIT + depreciation (if both present)
                ebitda = None
                if ebit is not None and depr is not None:
                    ebitda = ebit + depr

                fs_params = {
                    "orgnr": orgnr,
                    "year": year,
                    "account_view": ACCOUNT_VIEW,
                    "revenue": revenue,
                    "ebitda": ebitda,
                    "ebit": ebit,
                    "assets": assets,
                    "equity": equity,
                    "cash": cash,
                    "source": SOURCE_NAME,
                    "fetched_at_utc": fetched_at,
                }
                conn.execute(text(MERGE_FIN_STATEMENT), fs_params)

                # Optional: also write into proff_financial_item core codes
                for field, code in FIELD_TO_PROFF_CODE.items():
                    val = metrics.get(field)
                    if val is None:
                        continue
                    fin_item_params = {
                        "orgnr": orgnr,
                        "fiscal_year": year,
                        "account_view": ACCOUNT_VIEW,
                        "code": code,
                        "value": val,
                        "currency": "NOK",   # if sheet provides, map it; else assume NOK
                        "unit": None,
                        "fetched_at_utc": fetched_at,
                        "source": SOURCE_NAME,
                    }
                    conn.execute(text(MERGE_PROFF_FIN_ITEM), fin_item_params)

        # --- Sheet 2: Kontaktpersoner ---
        for r in contact_rows:
            orgnr = normalize_orgnr(r.get("Orgnr"))
            if not orgnr:
                continue

            params = {
                "orgnr": orgnr,
                "company_name": (str(r.get("Juridisk selskapsnavn")).strip() if r.get("Juridisk selskapsnavn") else None),
                "person_name": (str(r.get("Navn")).strip() if r.get("Navn") else None),
                "role": (str(r.get("Rolle")).strip() if r.get("Rolle") else None),
                "started_date": parse_date(r.get("Tiltrådt")),
                "phone": (str(r.get("Telefon")).strip() if r.get("Telefon") else None),
                "email": (str(r.get("E-post")).strip() if r.get("E-post") else None),
                "postal_address": (str(r.get("Gate-/postboksadresse (postadresse)")).strip() if r.get("Gate-/postboksadresse (postadresse)") else None),
                "postal_postnr": (str(r.get("Postnr (postadresse)")).strip() if r.get("Postnr (postadresse)") else None),
                "postal_city": (str(r.get("Poststed (postadresse)")).strip() if r.get("Poststed (postadresse)") else None),
                "business_address": (str(r.get("Gateadresse (forretningsadresse)")).strip() if r.get("Gateadresse (forretningsadresse)") else None),
                "business_postnr": (str(r.get("Postnr (forretningsadresse)")).strip() if r.get("Postnr (forretningsadresse)") else None),
                "business_city": (str(r.get("Poststed (forretningsadresse)")).strip() if r.get("Poststed (forretningsadresse)") else None),
                "revenue": parse_number(r.get("Driftsinntekter")),
                "employees": parse_int(r.get("Antall ansatte")),
                "source_file": source_file,
            }

            # Minimal validation
            if not params["person_name"]:
                continue

            conn.execute(text(MERGE_CONTACT), params)


def move_to_imported(path: Path) -> None:
    IMPORTED_DIR.mkdir(parents=True, exist_ok=True)
    ts = now_utc().strftime("%Y%m%d_%H%M%S")
    dest = IMPORTED_DIR / f"{path.stem}__imported_{ts}{path.suffix}"
    shutil.move(str(path), str(dest))


def main():
    engine = make_engine()
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    IMPORTED_DIR.mkdir(parents=True, exist_ok=True)

    files = sorted([p for p in IMPORT_DIR.glob("*.xlsx") if p.is_file()])
    if not files:
        print(f"No .xlsx files found in {IMPORT_DIR}")
        return

    for f in files:
        print(f"Importing: {f.name}")
        try:
            import_one_file(engine, f)
            move_to_imported(f)
            print(f"Imported and moved: {f.name}")
        except Exception as e:
            print(f"FAILED: {f.name} -> {e}")
            # Do not move file on failure
            continue


if __name__ == "__main__":
    main()
